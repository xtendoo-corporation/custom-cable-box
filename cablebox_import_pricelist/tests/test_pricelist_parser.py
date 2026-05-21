import importlib.util
import io
import pathlib
import unittest

from openpyxl import Workbook


MODULE_PATH = (
    pathlib.Path(__file__).resolve().parents[1]
    / "wizard"
    / "pricelist_parser.py"
)
spec = importlib.util.spec_from_file_location("pricelist_parser", MODULE_PATH)
assert spec is not None
assert spec.loader is not None
pricelist_parser = importlib.util.module_from_spec(spec)
spec.loader.exec_module(pricelist_parser)
PricelistWorkbookParser = pricelist_parser.PricelistWorkbookParser


class TestPricelistWorkbookParser(unittest.TestCase):
    def _build_workbook_bytes(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "TARIFA PAL"
        sheet.append(
            [
                "SIGLA MARCA",
                "CÓDIGO DE PRODUCTO",
                "CÓDIGO EAN",
                "DESCRIPCIÓN DEL PRODUCTO",
                "CANTIDAD DE CARTÓN",
                "CANTIDAD MÚLTIPLE",
                "CANTIDAD MÍNIMA",
                "CANTIDAD MÁXIMA",
                "PLAZOS DE ENTREGA",
                "PRECIO AL DISTRIBUIDOR",
                "PRECIO AL PÚBLICO",
                "MULTIPLICADOR PRECIO",
                "CÓDIGO DE MONEDA",
            ]
        )
        sheet.append(
            [
                "PAL",
                "BVFB1840SB",
                "ECO01744",
                "Producto A",
                12,
                1,
                1,
                999999,
                "4",
                "81,30",
                "91,30",
                1,
                "EUR",
            ]
        )
        sheet.append(
            [
                "PAL",
                "BVBW0840SB",
                "ECO02892",
                "Producto B",
                1,
                1,
                1,
                999999,
                "D",
                63.30,
                77.30,
                1,
                "EUR",
            ]
        )
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def test_parse_distributor_price(self):
        parser = PricelistWorkbookParser(price_field="distributor_price")
        sheets = parser.parse_workbook(self._build_workbook_bytes())

        self.assertEqual(len(sheets), 1)
        self.assertEqual(sheets[0]["name"], "TARIFA PAL")
        self.assertEqual(sheets[0]["currency_code"], "EUR")
        self.assertEqual(sheets[0]["rows"][0]["product_code"], "BVFB1840SB")
        self.assertEqual(sheets[0]["rows"][0]["ean"], "ECO01744")
        self.assertAlmostEqual(sheets[0]["rows"][0]["price"], 81.30)

    def test_parse_public_price(self):
        parser = PricelistWorkbookParser(price_field="public_price")
        sheets = parser.parse_workbook(self._build_workbook_bytes())

        self.assertAlmostEqual(sheets[0]["rows"][0]["price"], 91.30)
        self.assertAlmostEqual(sheets[0]["rows"][1]["price"], 77.30)

    def test_only_first_sheet_is_processed(self):
        workbook = Workbook()
        first_sheet = workbook.active
        first_sheet.title = "TARIFA 1"
        first_sheet.append(
            [
                "SIGLA MARCA",
                "CÓDIGO DE PRODUCTO",
                "CÓDIGO EAN",
                "DESCRIPCIÓN DEL PRODUCTO",
                "CANTIDAD DE CARTÓN",
                "CANTIDAD MÚLTIPLE",
                "CANTIDAD MÍNIMA",
                "CANTIDAD MÁXIMA",
                "PLAZOS DE ENTREGA",
                "PRECIO AL DISTRIBUIDOR",
                "PRECIO AL PÚBLICO",
                "MULTIPLICADOR PRECIO",
                "CÓDIGO DE MONEDA",
            ]
        )
        first_sheet.append(
            [
                "PAL",
                "PRIMERA",
                "EAN1",
                "Producto primera hoja",
                1,
                1,
                1,
                999999,
                "D",
                10,
                20,
                1,
                "EUR",
            ]
        )

        second_sheet = workbook.create_sheet("TARIFA 2")
        second_sheet.append(
            [
                "SIGLA MARCA",
                "CÓDIGO DE PRODUCTO",
                "CÓDIGO EAN",
                "DESCRIPCIÓN DEL PRODUCTO",
                "CANTIDAD DE CARTÓN",
                "CANTIDAD MÚLTIPLE",
                "CANTIDAD MÍNIMA",
                "CANTIDAD MÁXIMA",
                "PLAZOS DE ENTREGA",
                "PRECIO AL DISTRIBUIDOR",
                "PRECIO AL PÚBLICO",
                "MULTIPLICADOR PRECIO",
                "CÓDIGO DE MONEDA",
            ]
        )
        second_sheet.append(
            [
                "PAL",
                "SEGUNDA",
                "EAN2",
                "Producto segunda hoja",
                1,
                1,
                1,
                999999,
                "D",
                30,
                40,
                1,
                "EUR",
            ]
        )

        stream = io.BytesIO()
        workbook.save(stream)

        parser = PricelistWorkbookParser(price_field="distributor_price")
        sheets = parser.parse_workbook(stream.getvalue())

        self.assertEqual(len(sheets), 1)
        self.assertEqual(sheets[0]["name"], "TARIFA 1")
        self.assertEqual(len(sheets[0]["rows"]), 1)
        self.assertEqual(sheets[0]["rows"][0]["product_code"], "PRIMERA")

    def test_invalid_workbook_without_rows(self):
        workbook = Workbook()
        sheet = workbook.active
        sheet.title = "VACIA"
        sheet.append([
            "SIGLA MARCA",
            "CÓDIGO DE PRODUCTO",
            "CÓDIGO EAN",
            "DESCRIPCIÓN DEL PRODUCTO",
            "CANTIDAD DE CARTÓN",
            "CANTIDAD MÚLTIPLE",
            "CANTIDAD MÍNIMA",
            "CANTIDAD MÁXIMA",
            "PLAZOS DE ENTREGA",
            "PRECIO AL DISTRIBUIDOR",
            "PRECIO AL PÚBLICO",
            "MULTIPLICADOR PRECIO",
            "CÓDIGO DE MONEDA",
        ])
        stream = io.BytesIO()
        workbook.save(stream)

        parser = PricelistWorkbookParser()
        with self.assertRaises(ValueError):
            parser.parse_workbook(stream.getvalue())


if __name__ == "__main__":
    unittest.main()


