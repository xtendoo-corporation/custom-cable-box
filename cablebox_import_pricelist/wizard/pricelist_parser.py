import io

import openpyxl


class PricelistWorkbookParser:
    PRODUCT_CODE_COLUMN = 1
    EAN_COLUMN = 2
    DESCRIPTION_COLUMN = 3
    DISTRIBUTOR_PRICE_COLUMN = 9
    PUBLIC_PRICE_COLUMN = 10
    CURRENCY_CODE_COLUMN = 12
    MAX_USED_COLUMN = CURRENCY_CODE_COLUMN + 1
    DATA_START_ROW = 2

    def __init__(self, price_field="distributor_price"):
        if price_field not in {"distributor_price", "public_price"}:
            raise ValueError("Campo de precio no soportado.")
        self.price_field = price_field

    @staticmethod
    def normalize_text(value):
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    @staticmethod
    def normalize_price(value):
        if value in (None, ""):
            return None
        if isinstance(value, (int, float)):
            return float(value)

        text = str(value).strip()
        if not text:
            return None

        text = text.replace("€", "").replace(" ", "")
        if "," in text and "." in text:
            if text.rfind(",") > text.rfind("."):
                text = text.replace(".", "").replace(",", ".")
            else:
                text = text.replace(",", "")
        elif "," in text:
            text = text.replace(".", "").replace(",", ".")

        return float(text)

    def parse_workbook(self, file_content):
        workbook = openpyxl.load_workbook(
            io.BytesIO(file_content), data_only=True, read_only=True
        )
        if not workbook.worksheets:
            raise ValueError(
                "El Excel no contiene hojas para importar tarifas."
            )

        sheet = workbook.worksheets[0]
        parsed_sheet = self._parse_sheet(sheet)

        if not parsed_sheet:
            raise ValueError(
                "La primera hoja del Excel no contiene filas válidas para importar tarifas."
            )

        return [parsed_sheet]

    def _parse_sheet(self, sheet):
        if sheet.max_row < self.DATA_START_ROW:
            return None

        rows = []
        currencies = []
        price_column = self._get_price_column()
        for row_number, row_values in enumerate(
            sheet.iter_rows(
                min_row=self.DATA_START_ROW,
                max_col=self.MAX_USED_COLUMN,
                values_only=True,
            ),
            start=self.DATA_START_ROW,
        ):
            if not any(value not in (None, "") for value in row_values):
                continue

            product_code = self.normalize_text(self._get_row_value(row_values, self.PRODUCT_CODE_COLUMN))
            ean = self.normalize_text(self._get_row_value(row_values, self.EAN_COLUMN))
            description = self.normalize_text(
                self._get_row_value(row_values, self.DESCRIPTION_COLUMN)
            )
            currency_code = self.normalize_text(
                self._get_row_value(row_values, self.CURRENCY_CODE_COLUMN)
            ).upper()
            price = self.normalize_price(self._get_row_value(row_values, price_column))

            if not any([product_code, ean, description, currency_code, price is not None]):
                continue

            parsed_row = {
                "excel_row": row_number,
                "product_code": product_code,
                "ean": ean,
                "description": description,
                "currency_code": currency_code,
                "price": price,
            }
            if parsed_row["currency_code"]:
                currencies.append(parsed_row["currency_code"])
            rows.append(parsed_row)

        if not rows:
            return None

        sheet_currency = ""
        if currencies:
            sheet_currency = currencies[0]

        return {
            "name": sheet.title.strip() or "Tarifa sin nombre",
            "currency_code": sheet_currency,
            "rows": rows,
        }

    def _get_price_column(self):
        if self.price_field == "distributor_price":
            return self.DISTRIBUTOR_PRICE_COLUMN
        return self.PUBLIC_PRICE_COLUMN

    @staticmethod
    def _get_row_value(row_values, column_index):
        if column_index >= len(row_values):
            return None
        return row_values[column_index]

