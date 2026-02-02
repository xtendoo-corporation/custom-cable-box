# License AGPL-3.0 or later (http://www.gnu.org/licenses/agpl.html).

import io
import logging

_logger = logging.getLogger(__name__)

try:
    import openpyxl
except ImportError:
    openpyxl = None
    _logger.warning(
        "openpyxl library not found. Please install it with: pip install openpyxl"
    )


def parse_xlsx_products(file_data):
    """
    Parsea un archivo XLSX de productos Palazzoli y devuelve una lista de diccionarios.

    Args:
        file_data: Bytes del archivo XLSX

    Returns:
        list: Lista de diccionarios con los datos de productos
    """
    if not openpyxl:
        raise ImportError(
            "La librería openpyxl no está instalada. Ejecute: pip install openpyxl"
        )

    workbook = openpyxl.load_workbook(io.BytesIO(file_data), read_only=True)

    # Usar la primera hoja
    sheet_names = workbook.sheetnames
    sheet = workbook[sheet_names[0]]

    # Mapeo de columnas esperadas
    headers = {}
    header_row = None

    # Buscar fila de encabezados
    for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
        if row and row[0] and "SIGLA MARCA" in str(row).upper():
            header_row = row_idx
            for col_idx, cell in enumerate(row):
                if cell:
                    headers[str(cell).strip().upper()] = col_idx
            break

    if not header_row:
        # Buscar encabezados alternativos
        for row_idx, row in enumerate(sheet.iter_rows(values_only=True), start=1):
            if row and any(
                "CÓDIGO" in str(cell).upper() if cell else False for cell in row
            ):
                header_row = row_idx
                for col_idx, cell in enumerate(row):
                    if cell:
                        headers[str(cell).strip().upper()] = col_idx
                break

    if not headers:
        workbook.close()
        raise ValueError("No se encontraron encabezados válidos en el archivo.")

    # Mapeo de columnas del archivo a campos de Odoo
    column_mapping = {
        "CÓDIGO DE PRODUCTO": "default_code",
        "CÓDIGO EAN": "barcode",
        "DESCRIPCIÓN DEL PRODUCTO": "name",
        "BREVE DESCRIPCIÓN (MÁXIMO 43 CARACTERES)": "description_short",
        "PRECIO AL DISTRIBUIDOR": "standard_price",
        "PRECIO BRUTO DE LISTA": "list_price",
        "PESO BRUTO (KG)": "weight",
        "CANTIDAD DE CARTÓN": "qty_carton",
        "CANTIDAD MÍNIMA": "qty_min",
        "FAMILIA DE DESCUENTO": "discount_family",
        "FAMILIA DE PRODUCTO": "product_family",
        "PLAZOS DE ENTREGA": "delivery_lead_time",
    }

    # Encontrar índices de columnas
    col_indices = {}
    for xlsx_col, odoo_field in column_mapping.items():
        for header, idx in headers.items():
            if xlsx_col in header:
                col_indices[odoo_field] = idx
                break

    print(f"=== Headers encontrados en Excel: {list(headers.keys())}")
    print(f"=== Columnas mapeadas: {list(col_indices.keys())}")

    if "default_code" not in col_indices:
        workbook.close()
        raise ValueError(
            "No se encontró la columna 'CÓDIGO DE PRODUCTO' en el archivo."
        )

    products = []

    # Iterar sobre las filas de datos
    for row_idx, row in enumerate(
        sheet.iter_rows(min_row=header_row + 1, values_only=True), start=header_row + 1
    ):
        if not row or not row[col_indices.get("default_code", 0)]:
            continue

        default_code = str(row[col_indices["default_code"]]).strip()
        if not default_code or default_code == "None":
            continue

        # Preparar valores del producto
        product_data = {
            "default_code": default_code,
            "row_number": row_idx,
        }

        # Nombre del producto
        if "name" in col_indices and row[col_indices["name"]]:
            product_data["name"] = str(row[col_indices["name"]]).strip()
        else:
            product_data["name"] = default_code

        # Código de barras
        if "barcode" in col_indices and row[col_indices["barcode"]]:
            barcode = str(row[col_indices["barcode"]]).strip()
            if barcode and barcode != "None":
                product_data["barcode"] = barcode

        # Precio de coste
        if "standard_price" in col_indices and row[col_indices["standard_price"]]:
            try:
                product_data["standard_price"] = float(
                    row[col_indices["standard_price"]]
                )
            except (ValueError, TypeError):
                pass

        # Precio de venta
        if "list_price" in col_indices and row[col_indices["list_price"]]:
            try:
                product_data["list_price"] = float(row[col_indices["list_price"]])
            except (ValueError, TypeError):
                pass

        # Peso
        if "weight" in col_indices and row[col_indices["weight"]]:
            try:
                product_data["weight"] = float(row[col_indices["weight"]])
            except (ValueError, TypeError):
                pass

        # Familia de Descuento (categoría padre)
        if "discount_family" in col_indices and row[col_indices["discount_family"]]:
            discount_family = str(row[col_indices["discount_family"]]).strip()
            if discount_family and discount_family != "None":
                product_data["discount_family"] = discount_family

        # Familia de Producto (subcategoría)
        if "product_family" in col_indices and row[col_indices["product_family"]]:
            product_family = str(row[col_indices["product_family"]]).strip()
            if product_family and product_family != "None":
                product_data["product_family"] = product_family

        # Plazo de entrega (sale_delay) - 'd' significa 25 días
        if (
            "delivery_lead_time" in col_indices
            and row[col_indices["delivery_lead_time"]]
        ):
            lead_time_value = (
                str(row[col_indices["delivery_lead_time"]]).strip().lower()
            )
            if lead_time_value == "d":
                product_data["sale_delay"] = 25
            elif lead_time_value and lead_time_value != "none":
                try:
                    product_data["sale_delay"] = int(float(lead_time_value))
                except (ValueError, TypeError):
                    pass

        products.append(product_data)

    workbook.close()
    return products


def get_or_create_category(env, discount_family, product_family):
    """
    Busca o crea una categoría jerárquica.
    FAMILIA DE DESCUENTO es el padre, FAMILIA DE PRODUCTO es el hijo.

    Args:
        env: Odoo environment
        discount_family: Nombre de la categoría padre (ej: 'LUX')
        product_family: Nombre de la subcategoría (ej: 'SI')

    Returns:
        int: ID de la categoría hija (ej: LUX/SI)
    """
    Category = env["product.category"]

    # Buscar o crear categoría padre
    parent_category = Category.search(
        [("name", "=", discount_family), ("parent_id", "=", False)], limit=1
    )
    if not parent_category:
        parent_category = Category.create({"name": discount_family})
        _logger.info(f"Categoría padre creada: {discount_family}")

    # Buscar o crear subcategoría
    child_category = Category.search(
        [("name", "=", product_family), ("parent_id", "=", parent_category.id)], limit=1
    )
    if not child_category:
        child_category = Category.create(
            {
                "name": product_family,
                "parent_id": parent_category.id,
            }
        )
        _logger.info(f"Subcategoría creada: {discount_family}/{product_family}")

    return child_category.id


def import_products_to_odoo(env, products, update_existing=True):
    """
    Importa los productos parseados a Odoo.

    Args:
        env: Odoo environment
        products: Lista de diccionarios con datos de productos
        update_existing: Si es True, actualiza productos existentes

    Returns:
        dict: Estadísticas de la importación
    """
    Product = env["product.template"]
    created_count = 0
    updated_count = 0
    skipped_count = 0
    errors = []

    # Cache para categorías ya procesadas
    category_cache = {}

    for product_data in products:
        try:
            row_number = product_data.pop("row_number", 0)
            default_code = product_data["default_code"]

            # Extraer familias antes de crear vals
            discount_family = product_data.pop("discount_family", None)
            product_family = product_data.pop("product_family", None)

            # Preparar valores base (en Odoo 18, 'consu' es el tipo para productos almacenables)
            vals = {
                "default_code": default_code,
                "type": "consu",
                "sale_ok": True,
                "purchase_ok": True,
            }
            vals.update(product_data)

            # Asignar categoría jerárquica si hay ambas familias
            if discount_family and product_family:
                cache_key = f"{discount_family}|{product_family}"
                if cache_key not in category_cache:
                    category_cache[cache_key] = get_or_create_category(
                        env, discount_family, product_family
                    )
                vals["categ_id"] = category_cache[cache_key]

            # Buscar producto existente
            existing_product = Product.search(
                [("default_code", "=", default_code)], limit=1
            )

            if existing_product:
                if update_existing:
                    existing_product.write(vals)
                    updated_count += 1
                else:
                    skipped_count += 1
            else:
                Product.create(vals)
                created_count += 1

        except Exception as e:
            errors.append(f"Fila {row_number}: {str(e)}")
            _logger.error(f"Error en fila {row_number}: {str(e)}")

    return {
        "created": created_count,
        "updated": updated_count,
        "skipped": skipped_count,
        "errors": errors,
    }
